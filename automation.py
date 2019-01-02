
from  constants import *
import glob, os
from xlrd import open_workbook
import win32com.client as w3c
from openpyxl import *

team_arr = []
raw_file_path = ""

base_path = "C:\\workspace\\python\\evaluation_automation\\"
def get_files(team_kind):
    # 폴더에서 파일 읽기
    data_path = base_path + "data\\"
    for filename in os.listdir(data_path):
        raw_file_path = data_path + filename
        raw_wb = load_workbook(raw_file_path)
        raw_sheet = raw_wb.worksheets[0]

        # 부서명, 인명 읽기
        if(team_kind == team):
            team_name = raw_sheet["c4"].value

        else:
            team_name = raw_sheet["f4"].value
        person_name =  raw_sheet["o4"].value

        # 파일의 부서명에 해당하는 엑셀파일이 있는지 확인
        single_path = team_name + ".xlsx"
        if team_name in team_arr :
            # 있다면 해당 파일 불러오기
            print("team file exists" + team_name + " " + person_name)
            team_wb = load_workbook(base_path + single_path)
        else:
            # 없다면 파일 만들기
            print(team_name + " " + person_name)
            team_wb = Workbook()
            team_wb.save(single_path)

            team_arr.append(team_name)

        # 시트 추가하기
        copy_sheet(raw_file_path, single_path, person_name)

def copy_sheet(raw_file_path, single_path, person_name):
    # xl = Dispatch("Excel.Application")
    excel = w3c.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = True
    wb1 = excel.Workbooks.Open(raw_file_path)
    wb2 = excel.Workbooks.Open(base_path + single_path)
    ws1 = wb1.Worksheets(1)
    ws1.Name = person_name
    ws1.Copy(Before=wb2.Worksheets(1))

    wb1.Close(SaveChanges=True)
    wb2.Close(SaveChanges=True)

    excel.Quit()

# 여기에 team과 sub_team을 넣어서 호출하면 된다
get_files(sub_team)
